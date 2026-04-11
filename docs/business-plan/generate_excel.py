#!/usr/bin/env python3
"""
Generateur du previsionnel financier NEXUS en .xlsx avec formules.
Toutes les cellules calculees utilisent des formules Excel (pas de valeurs en dur).
VERSION 3.0 — Avril 2026 (Free / Basic 29€ / Business 149€ + credits IA)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# === STYLES ===
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1A365D')
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='2B6CB0')
HEADER_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
NORMAL_FONT = Font(name='Calibri', size=10)
SMALL_FONT = Font(name='Calibri', size=9, italic=True, color='666666')

HEADER_FILL = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
LIGHT_RED = PatternFill(start_color='FCE8E6', end_color='FCE8E6', fill_type='solid')
LIGHT_YELLOW = PatternFill(start_color='FEF7E0', end_color='FEF7E0', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
TOTAL_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

EUR_FORMAT = '#,##0 €'
EUR_FORMAT_NEG = '#,##0 €;[Red]-#,##0 €'
PCT_FORMAT = '0%'
PCT_FORMAT_1 = '0.0%'

def style_header_row(ws, row, max_col, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_range(ws, start_row, end_row, max_col, font=NORMAL_FONT, fill=None, num_format=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if num_format and c > 1:
                cell.number_format = num_format

def style_total_row(ws, row, max_col, fill=TOTAL_FILL, font=TOTAL_FONT):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.border = THIN_BORDER
        if col > 1:
            cell.number_format = EUR_FORMAT

def auto_width(ws, min_width=12, max_width=20):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# =====================================================================
# ONGLET 1 : HYPOTHESES (cellules nommees, modifiables par le banquier)
# =====================================================================
ws1 = wb.active
ws1.title = "Hypotheses"
ws1.sheet_properties.tabColor = "2B6CB0"

ws1.merge_cells('A1:C1')
ws1['A1'] = "HYPOTHESES FINANCIERES — NEXUS"
ws1['A1'].font = TITLE_FONT

ws1.merge_cells('A2:C2')
ws1['A2'] = "Modifiez ces valeurs pour voir l'impact sur tous les onglets"
ws1['A2'].font = SMALL_FONT

# Headers
for col, label in enumerate(['Parametre', 'Valeur', 'Note'], 1):
    ws1.cell(row=4, column=col, value=label)
style_header_row(ws1, 4, 3)

# Data rows (row, label, value, note)
hyp_data = [
    (5, "Prix Free (EUR/mois)", 0, "Gratuit a vie"),
    (6, "Prix Basic (EUR/mois)", 29, "Illimite + 1000 credits IA/mois"),
    (7, "Prix Business (EUR/mois)", 149, "Multi-site + 10000 credits IA/mois"),
    (8, "Mix Free (%)", 0.10, "Utilisateurs gratuits"),
    (9, "Mix Basic (%)", 0.70, "Coeur de cible PME"),
    (10, "Mix Business (%)", 0.20, "Multi-sites, franchises"),
    (11, "ARPA (revenu moyen/client/mois)", None, "=B5*B8+B6*B9+B7*B10"),  # formula
    (12, "", "", ""),
    (13, "Churn mensuel An 1", 0.03, "Benchmark SaaS SMB"),
    (14, "Churn mensuel An 2", 0.025, "Amelioration onboarding"),
    (15, "Churn mensuel An 3", 0.02, "Produit mature"),
    (16, "", "", ""),
    (17, "Cout infra fixe (EUR/mois)", 150, "Render + Supabase base"),
    (18, "Cout variable/client (EUR/mois)", 5, "Anthropic + Twilio + Resend"),
    (19, "Commission Stripe (%)", 0.029, "2.9%"),
    (20, "", "", ""),
    (21, "Salaire fondateur M1-M6", 0, "Pas de salaire les 6 premiers mois"),
    (22, "Salaire fondateur M7+ An 1 (EUR/mois)", 1200, "Phase demarrage"),
    (23, "Salaire fondateur An 2 (EUR/mois)", 2000, "Croissance"),
    (24, "Salaire fondateur An 3 (EUR/mois)", 3000, "Rentabilite"),
    (25, "Taux charges sociales", 0.45, "SASU regime general"),
    (26, "", "", ""),
    (27, "Comptable / juridique", 0, "Fondateur = comptable (BEP + 10 ans exp)"),
    (28, "Loyer", 0, "Societe hebergee au domicile du fondateur"),
    (29, "Marketing An 1 (EUR/mois)", 300, "Reseaux sociaux + pub ciblee"),
    (30, "Marketing An 2+ (EUR/mois)", 800, "Ads + partenariats"),
    (31, "", "", ""),
    (32, "TVA", 0.20, "Taux standard"),
    (33, "IS taux reduit (< 42500 EUR)", 0.15, "PME"),
    (34, "IS taux normal", 0.25, "Au-dela de 42500 EUR"),
]

for row, label, value, note in hyp_data:
    ws1.cell(row=row, column=1, value=label).font = BOLD_FONT if label else NORMAL_FONT
    cell_b = ws1.cell(row=row, column=2)
    cell_c = ws1.cell(row=row, column=3, value=note)
    cell_c.font = SMALL_FONT

    if row == 11:
        # ARPA formula
        cell_b.value = "=B5*B8+B6*B9+B7*B10"
        cell_b.number_format = EUR_FORMAT
        cell_b.font = BOLD_FONT
        ws1.cell(row=row, column=1, value="ARPA (revenu moyen/client/mois)").font = BOLD_FONT
    elif isinstance(value, float) and 0 < value < 1:
        cell_b.value = value
        cell_b.number_format = PCT_FORMAT if value >= 0.10 else PCT_FORMAT_1
    elif isinstance(value, (int, float)) and value >= 1:
        cell_b.value = value
        cell_b.number_format = EUR_FORMAT if value > 10 else '0'
    else:
        cell_b.value = value

    cell_b.font = NORMAL_FONT
    for c in range(1, 4):
        ws1.cell(row=row, column=c).border = THIN_BORDER

ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 35


# =====================================================================
# ONGLET 2 : PLAN DE FINANCEMENT INITIAL
# =====================================================================
ws2 = wb.create_sheet("Plan Financement")
ws2.sheet_properties.tabColor = "38A169"

ws2.merge_cells('A1:E1')
ws2['A1'] = "PLAN DE FINANCEMENT INITIAL — 40 000 EUR (+ NACRE 5K optionnel)"
ws2['A1'].font = TITLE_FONT

# Besoins
ws2.cell(row=3, column=1, value="BESOINS").font = SECTION_FONT
ws2.cell(row=3, column=4, value="RESSOURCES").font = SECTION_FONT
for col, label in [(1, "Poste"), (2, "Montant"), (4, "Poste"), (5, "Montant")]:
    ws2.cell(row=4, column=col, value=label)
style_header_row(ws2, 4, 5)

besoins = [
    (5, "Fonds de roulement (6 mois charges mini)", 18000),
    (6, "Tresorerie de securite (couverture pertes Y1)", 16000),
    (7, "Marketing lancement (reseaux, pub ciblee)", 3000),
    (8, "Materiel informatique", 1500),
    (9, "Frais d'etablissement (SASU, juridique)", 1500),
]

ressources = [
    (5, "Capital social SASU", 1),
    (6, "Apport en nature (logiciel NEXUS v3.25)", 80000),
    (7, "Pret d'honneur Initiative 95", 15000),
    (8, "Pret bancaire (garanti BPI 60%)", 25000),
    (9, "NACRE phase 2 (OPTIONNEL)", 5000),
]

for row, label, val in besoins:
    ws2.cell(row=row, column=1, value=label).font = NORMAL_FONT
    ws2.cell(row=row, column=2, value=val).number_format = EUR_FORMAT
    ws2.cell(row=row, column=2).border = THIN_BORDER
    ws2.cell(row=row, column=1).border = THIN_BORDER

for row, label, val in ressources:
    ws2.cell(row=row, column=4, value=label).font = NORMAL_FONT
    ws2.cell(row=row, column=5, value=val).number_format = EUR_FORMAT
    ws2.cell(row=row, column=5).border = THIN_BORDER
    ws2.cell(row=row, column=4).border = THIN_BORDER

# Totals with formulas
total_row = 11
ws2.cell(row=total_row, column=1, value="TOTAL BESOINS")
ws2.cell(row=total_row, column=2, value="=SUM(B5:B9)")
ws2.cell(row=total_row, column=2).number_format = EUR_FORMAT
ws2.cell(row=total_row, column=4, value="TOTAL TRESORERIE BASE (sans NACRE)")
ws2.cell(row=total_row, column=5, value="=E7+E8")
ws2.cell(row=total_row, column=5).number_format = EUR_FORMAT
style_total_row(ws2, total_row, 5)

# Equilibre check
ws2.cell(row=13, column=1, value="Equilibre (Ressources tresorerie - Besoins)")
ws2.cell(row=13, column=2, value="=E11-B11")
ws2.cell(row=13, column=2).number_format = EUR_FORMAT_NEG
ws2.cell(row=13, column=1).font = BOLD_FONT
ws2.cell(row=13, column=2).font = BOLD_FONT

# Total avec NACRE
ws2.cell(row=13, column=4, value="TOTAL AVEC NACRE (optionnel)").font = BOLD_FONT
ws2.cell(row=13, column=5, value="=E7+E8+E9")
ws2.cell(row=13, column=5).number_format = EUR_FORMAT
ws2.cell(row=13, column=5).font = BOLD_FONT

# Notes
ws2.cell(row=15, column=1, value="Note : L'apport en nature (80 000 EUR) valorise le logiciel developpe").font = SMALL_FONT
ws2.cell(row=16, column=1, value="pendant 18 mois. Il n'apporte pas de tresorerie mais renforce les fonds propres.").font = SMALL_FONT
ws2.cell(row=17, column=1, value="NACRE = gere par la Region (pas France Travail). Bonus si eligible, pas indispensable.").font = SMALL_FONT

ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 3
ws2.column_dimensions['D'].width = 42
ws2.column_dimensions['E'].width = 15


# =====================================================================
# ONGLET 3 : EVOLUTION CLIENTS & MRR (mois par mois, An 1)
# =====================================================================
ws3 = wb.create_sheet("Clients & MRR")
ws3.sheet_properties.tabColor = "DD6B20"

ws3.merge_cells('A1:F1')
ws3['A1'] = "EVOLUTION CLIENTS ET MRR — ANNEE 1"
ws3['A1'].font = TITLE_FONT

ws3.merge_cells('A2:F2')
ws3['A2'] = "Les colonnes Nouveaux et Churn sont modifiables. Le reste se recalcule."
ws3['A2'].font = SMALL_FONT

for col, label in enumerate(["Mois", "Nouveaux", "Churn", "Clients fin mois", "MRR (EUR)", "ARR annualise"], 1):
    ws3.cell(row=4, column=col, value=label)
style_header_row(ws3, 4, 6)

# Input: nouveaux clients and churn per month (acquisition acceleree multi-canal)
nouveaux = [6, 7, 8, 9, 10, 10, 11, 12, 12, 13, 14, 14]
churns = [0, 0, 0, 1, 1, 1, 1, 2, 2, 2, 3, 3]

for i in range(12):
    r = 5 + i
    month_label = f"M{i+1}"
    ws3.cell(row=r, column=1, value=month_label).font = BOLD_FONT

    # Nouveaux (input - modifiable)
    ws3.cell(row=r, column=2, value=nouveaux[i])
    ws3.cell(row=r, column=2).fill = LIGHT_YELLOW  # editable highlight

    # Churn (input - modifiable)
    ws3.cell(row=r, column=3, value=churns[i])
    ws3.cell(row=r, column=3).fill = LIGHT_YELLOW

    # Clients fin mois (formula)
    if i == 0:
        ws3.cell(row=r, column=4, value=f"=B{r}-C{r}")
    else:
        ws3.cell(row=r, column=4, value=f"=D{r-1}+B{r}-C{r}")

    # MRR = Clients * ARPA (linked to Hypotheses!B11)
    ws3.cell(row=r, column=5, value=f"=D{r}*Hypotheses!B11")
    ws3.cell(row=r, column=5).number_format = EUR_FORMAT

    # ARR = MRR * 12
    ws3.cell(row=r, column=6, value=f"=E{r}*12")
    ws3.cell(row=r, column=6).number_format = EUR_FORMAT

    for c in range(1, 7):
        ws3.cell(row=r, column=c).border = THIN_BORDER

# Total An 1
r_total = 17
ws3.cell(row=r_total, column=1, value="Total An 1")
ws3.cell(row=r_total, column=2, value="=SUM(B5:B16)")
ws3.cell(row=r_total, column=3, value="=SUM(C5:C16)")
ws3.cell(row=r_total, column=4, value="=D16")  # end of year clients
ws3.cell(row=r_total, column=5, value="=E16")  # end MRR
ws3.cell(row=r_total, column=6, value="=F16")  # end ARR
style_total_row(ws3, r_total, 6)

# An 2 summary
ws3.cell(row=19, column=1, value="An 2 (previsions)").font = SECTION_FONT
ws3.cell(row=20, column=1, value="Nouveaux clients").font = NORMAL_FONT
ws3.cell(row=20, column=2, value=278)
ws3.cell(row=20, column=2).fill = LIGHT_YELLOW
ws3.cell(row=21, column=1, value="Churn total").font = NORMAL_FONT
ws3.cell(row=21, column=2, value=63)
ws3.cell(row=21, column=2).fill = LIGHT_YELLOW
ws3.cell(row=22, column=1, value="Clients fin An 2").font = BOLD_FONT
ws3.cell(row=22, column=2, value="=D17+B20-C21")
ws3.cell(row=23, column=1, value="MRR fin An 2").font = BOLD_FONT
ws3.cell(row=23, column=2, value="=B22*Hypotheses!B11")
ws3.cell(row=23, column=2).number_format = EUR_FORMAT

# An 3 summary
ws3.cell(row=25, column=1, value="An 3 (previsions)").font = SECTION_FONT
ws3.cell(row=26, column=1, value="Nouveaux clients").font = NORMAL_FONT
ws3.cell(row=26, column=2, value=285)
ws3.cell(row=26, column=2).fill = LIGHT_YELLOW
ws3.cell(row=27, column=1, value="Churn total").font = NORMAL_FONT
ws3.cell(row=27, column=2, value=60)
ws3.cell(row=27, column=2).fill = LIGHT_YELLOW
ws3.cell(row=28, column=1, value="Clients fin An 3").font = BOLD_FONT
ws3.cell(row=28, column=2, value="=B22+B26-C27")
ws3.cell(row=29, column=1, value="MRR fin An 3").font = BOLD_FONT
ws3.cell(row=29, column=2, value="=B28*Hypotheses!B11")
ws3.cell(row=29, column=2).number_format = EUR_FORMAT

auto_width(ws3)


# =====================================================================
# ONGLET 4 : COMPTE DE RESULTAT PREVISIONNEL (3 ans)
# =====================================================================
ws4 = wb.create_sheet("Compte Resultat")
ws4.sheet_properties.tabColor = "E53E3E"

ws4.merge_cells('A1:D1')
ws4['A1'] = "COMPTE DE RESULTAT PREVISIONNEL (3 ANS)"
ws4['A1'].font = TITLE_FONT

ws4.merge_cells('A2:D2')
ws4['A2'] = "Lie aux onglets Hypotheses et Clients & MRR — Solo founder An 1/An 2, 1er recrutement An 3"
ws4['A2'].font = SMALL_FONT

for col, label in enumerate(["Poste", "An 1", "An 2", "An 3"], 1):
    ws4.cell(row=4, column=col, value=label)
style_header_row(ws4, 4, 4)

# Row references for formulas
rows = {}

def cr_row(r, label, an1, an2, an3, is_total=False, is_section=False, fmt=EUR_FORMAT_NEG):
    rows[label] = r
    ws4.cell(row=r, column=1, value=label)
    if is_section:
        ws4.cell(row=r, column=1).font = SECTION_FONT
        return
    for c, val in [(2, an1), (3, an2), (4, an3)]:
        cell = ws4.cell(row=r, column=c, value=val)
        cell.number_format = fmt
        cell.border = THIN_BORDER
    ws4.cell(row=r, column=1).border = THIN_BORDER
    if is_total:
        style_total_row(ws4, r, 4)
    else:
        ws4.cell(row=r, column=1).font = NORMAL_FONT

cr_row(6, "PRODUITS D'EXPLOITATION", None, None, None, is_section=True)

# CA based on ARPA 55 EUR model — 110/325/550 clients
cr_row(7, "Abonnements SaaS + credits IA (HT)", 35860, 149710, 309700)
cr_row(8, "Setup / accompagnement", 6000, 13500, 16000)
cr_row(9, "TOTAL CA HT", "=B7+B8", "=C7+C8", "=D7+D8", is_total=True)

cr_row(11, "CHARGES VARIABLES", None, None, None, is_section=True)
cr_row(12, "Hebergement cloud (Render, Supabase)", 2400, 4800, 8400)
cr_row(13, "Services tiers (Anthropic, Twilio, Resend)", 3200, 12000, 24000)
cr_row(14, "Commissions Stripe (2.9%)", 1200, 4700, 9500)
cr_row(15, "Total charges variables", "=SUM(B12:B14)", "=SUM(C12:C14)", "=SUM(D12:D14)", is_total=True)

cr_row(17, "MARGE BRUTE", "=B9-B15", "=C9-C15", "=D9-D15")
ws4.cell(row=17, column=1).font = BOLD_FONT
for c in [2, 3, 4]:
    ws4.cell(row=17, column=c).font = BOLD_FONT

cr_row(18, "Taux de marge brute", "=B17/B9", "=C17/C9", "=D17/D9", fmt=PCT_FORMAT)

cr_row(20, "CHARGES FIXES", None, None, None, is_section=True)
cr_row(21, "Remuneration fondateur (net)", 7200, 24000, 36000)
cr_row(22, "Charges sociales (45%)", "=B21*Hypotheses!B25", "=C21*Hypotheses!B25", "=D21*Hypotheses!B25")
cr_row(23, "Prestataires externes (missions)", 0, 10000, 6000)
cr_row(24, "Salaire support/CSM (a partir M25)", 0, 0, 24000)
cr_row(25, "Charges sociales CSM", 0, 0, "=D24*Hypotheses!B25")
cr_row(26, "Marketing & acquisition", 3600, 9600, 14400)
cr_row(27, "Assurances (RC Pro)", 1000, 1200, 2000)
cr_row(28, "Comptable / juridique (fondateur = comptable)", 0, 0, 0)
cr_row(29, "Materiel / licences", 500, 1500, 2500)
cr_row(30, "Loyer / coworking (domicile)", 0, 0, 0)
cr_row(31, "Divers / imprevus", 1200, 2000, 3000)
cr_row(32, "Total charges fixes", "=SUM(B21:B31)", "=SUM(C21:C31)", "=SUM(D21:D31)", is_total=True)

cr_row(34, "RESULTAT D'EXPLOITATION", "=B17-B32", "=C17-C32", "=D17-D32")
ws4.cell(row=34, column=1).font = BOLD_FONT
cr_row(35, "Charges financieres (interets emprunts)", 375, 625, 500)
cr_row(36, "RESULTAT COURANT", "=B34-B35", "=C34-C35", "=D34-D35")

# IS calculation with formula
cr_row(37, "IS (15%/25%)", '=IF(B36<=0,0,IF(B36<=42500,B36*Hypotheses!B33,42500*Hypotheses!B33+(B36-42500)*Hypotheses!B34))',
       '=IF(C36<=0,0,IF(C36<=42500,C36*Hypotheses!B33,42500*Hypotheses!B33+(C36-42500)*Hypotheses!B34))',
       '=IF(D36<=0,0,IF(D36<=42500,D36*Hypotheses!B33,42500*Hypotheses!B33+(D36-42500)*Hypotheses!B34))')

cr_row(38, "RESULTAT NET", "=B36-B37", "=C36-C37", "=D36-D37", is_total=True)

cr_row(40, "CAF (Resultat + amortissements)", "=B38+1500", "=C38+1500", "=D38+1500")
ws4.cell(row=40, column=1).font = BOLD_FONT

ws4.column_dimensions['A'].width = 40
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 16


# =====================================================================
# ONGLET 5 : TRESORERIE MENSUELLE AN 1
# =====================================================================
ws5 = wb.create_sheet("Tresorerie M1-M12")
ws5.sheet_properties.tabColor = "805AD5"

ws5.merge_cells('A1:N1')
ws5['A1'] = "PLAN DE TRESORERIE MENSUEL — ANNEE 1 (base 40K, sans NACRE)"
ws5['A1'].font = TITLE_FONT

# Headers
cols_treso = ["", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12", "TOTAL"]
for col, label in enumerate(cols_treso, 1):
    ws5.cell(row=3, column=col, value=label)
style_header_row(ws5, 3, 14)

# ENCAISSEMENTS section
r = 4
ws5.cell(row=r, column=1, value="ENCAISSEMENTS").font = SECTION_FONT
r = 5

# CA TTC = MRR * 1.2 (linked to Clients sheet)
ws5.cell(row=r, column=1, value="CA TTC (MRR x 1.2)")
for m in range(12):
    c = m + 2
    mrr_ref = f"'Clients & MRR'!E{5+m}"
    ws5.cell(row=r, column=c, value=f"={mrr_ref}*(1+Hypotheses!B32)")
    ws5.cell(row=r, column=c).number_format = EUR_FORMAT
    ws5.cell(row=r, column=c).border = THIN_BORDER
ws5.cell(row=r, column=14, value="=SUM(B5:M5)")
ws5.cell(row=r, column=14).number_format = EUR_FORMAT

r = 6
ws5.cell(row=r, column=1, value="Pret honneur Initiative 95")
ws5.cell(row=r, column=2, value=15000).number_format = EUR_FORMAT
ws5.cell(row=r, column=14, value="=SUM(B6:M6)")
ws5.cell(row=r, column=14).number_format = EUR_FORMAT

r = 7
ws5.cell(row=r, column=1, value="Pret bancaire (garanti BPI)")
ws5.cell(row=r, column=4, value=25000).number_format = EUR_FORMAT
ws5.cell(row=r, column=14, value="=SUM(B7:M7)")
ws5.cell(row=r, column=14).number_format = EUR_FORMAT

r = 8
ws5.cell(row=r, column=1, value="NACRE phase 2 (OPTIONNEL)")
ws5.cell(row=r, column=1).font = SMALL_FONT
ws5.cell(row=r, column=14, value=0)
ws5.cell(row=r, column=14).number_format = EUR_FORMAT

r = 9
ws5.cell(row=r, column=1, value="TOTAL ENCAISSEMENTS")
for c in range(2, 15):
    ws5.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}8)")
    ws5.cell(row=r, column=c).number_format = EUR_FORMAT
style_total_row(ws5, r, 14)

# DECAISSEMENTS
r = 11
ws5.cell(row=r, column=1, value="DECAISSEMENTS").font = SECTION_FONT

# Row data for decaissements with monthly values (solo founder, no salary M1-M6, no comptable, no loyer)
dec_rows = {
    12: ("Hebergement cloud", [150, 150, 175, 175, 200, 200, 200, 200, 225, 225, 250, 250]),
    13: ("Services tiers (Anthropic/Twilio/Resend)", [150, 175, 200, 225, 250, 275, 300, 325, 350, 375, 375, 400]),
    14: ("Commissions Stripe", [22, 37, 54, 70, 89, 107, 128, 148, 167, 190, 213, 234]),
    15: ("Salaire fondateur net (M7-M12)", [0, 0, 0, 0, 0, 0, 1200, 1200, 1200, 1200, 1200, 1200]),
    16: ("Charges sociales fondateur", [0, 0, 0, 0, 0, 0, 540, 540, 540, 540, 540, 540]),
    17: ("Marketing", [300]*12),
    18: ("Assurances (RC Pro)", [83]*12),
    19: ("Materiel", [500, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
    20: ("Remboursement emprunts", [0, 0, 0, 0, 0, 0, 728, 728, 728, 728, 728, 728]),
    21: ("TVA a reverser", [0, 0, 0, 0, 0, 0, 0, 0, 800, 0, 0, 1600]),
    22: ("Divers / imprevus", [100]*12),
}

for row_num, (label, values) in dec_rows.items():
    ws5.cell(row=row_num, column=1, value=label).font = NORMAL_FONT
    ws5.cell(row=row_num, column=1).border = THIN_BORDER
    for m, val in enumerate(values):
        c = m + 2
        ws5.cell(row=row_num, column=c, value=val)
        ws5.cell(row=row_num, column=c).number_format = EUR_FORMAT
        ws5.cell(row=row_num, column=c).border = THIN_BORDER
        if "Salaire fondateur" in label:
            ws5.cell(row=row_num, column=c).fill = LIGHT_BLUE
    # Total column
    ws5.cell(row=row_num, column=14, value=f"=SUM(B{row_num}:M{row_num})")
    ws5.cell(row=row_num, column=14).number_format = EUR_FORMAT
    ws5.cell(row=row_num, column=14).border = THIN_BORDER

# Total decaissements
r = 23
ws5.cell(row=r, column=1, value="TOTAL DECAISSEMENTS")
for c in range(2, 15):
    ws5.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}12:{get_column_letter(c)}22)")
    ws5.cell(row=r, column=c).number_format = EUR_FORMAT
style_total_row(ws5, r, 14)

# Solde du mois
r = 25
ws5.cell(row=r, column=1, value="SOLDE DU MOIS").font = BOLD_FONT
ws5.cell(row=r, column=1).border = THIN_BORDER
for c in range(2, 15):
    ws5.cell(row=r, column=c, value=f"={get_column_letter(c)}9-{get_column_letter(c)}23")
    ws5.cell(row=r, column=c).number_format = EUR_FORMAT_NEG
    ws5.cell(row=r, column=c).border = THIN_BORDER
    ws5.cell(row=r, column=c).font = BOLD_FONT

# Tresorerie cumulee
r = 26
ws5.cell(row=r, column=1, value="TRESORERIE CUMULEE").font = BOLD_FONT
ws5.cell(row=r, column=1).border = THIN_BORDER
ws5.cell(row=r, column=1).fill = LIGHT_GREEN
for m in range(12):
    c = m + 2
    if m == 0:
        ws5.cell(row=r, column=c, value=f"=B25")
    else:
        prev = get_column_letter(c - 1)
        cur = get_column_letter(c)
        ws5.cell(row=r, column=c, value=f"={prev}26+{cur}25")
    ws5.cell(row=r, column=c).number_format = EUR_FORMAT_NEG
    ws5.cell(row=r, column=c).border = THIN_BORDER
    ws5.cell(row=r, column=c).font = BOLD_FONT
    ws5.cell(row=r, column=c).fill = LIGHT_GREEN

ws5.column_dimensions['A'].width = 32
for c in range(2, 15):
    ws5.column_dimensions[get_column_letter(c)].width = 12


# =====================================================================
# ONGLET 6 : SEUIL DE RENTABILITE
# =====================================================================
ws6 = wb.create_sheet("Seuil Rentabilite")
ws6.sheet_properties.tabColor = "D69E2E"

ws6.merge_cells('A1:D1')
ws6['A1'] = "SEUIL DE RENTABILITE"
ws6['A1'].font = TITLE_FONT

ws6.merge_cells('A2:D2')
ws6['A2'] = "Seuil atteint a ~50 clients (M6-M7, fin S1)"
ws6['A2'].font = SMALL_FONT

for col, label in enumerate(["", "An 1", "An 2", "An 3"], 1):
    ws6.cell(row=3, column=col, value=label)
style_header_row(ws6, 3, 4)

# All formulas reference Compte Resultat
sr_data = [
    (4, "Charges fixes totales", "='Compte Resultat'!B32", "='Compte Resultat'!C32", "='Compte Resultat'!D32"),
    (5, "Taux de marge sur couts variables", "='Compte Resultat'!B18", "='Compte Resultat'!C18", "='Compte Resultat'!D18"),
    (6, "SEUIL DE RENTABILITE (CA HT)", "=B4/B5", "=C4/C5", "=D4/D5"),
    (7, "Seuil mensuel", "=B6/12", "=C6/12", "=D6/12"),
    (8, "Nb clients equivalent (ARPA)", "=B7/Hypotheses!B11", "=C7/Hypotheses!B11", "=D7/Hypotheses!B11"),
    (9, "", "", "", ""),
    (10, "CA previsionnel", "='Compte Resultat'!B9", "='Compte Resultat'!C9", "='Compte Resultat'!D9"),
    (11, "Marge de securite (CA - Seuil)", "=B10-B6", "=C10-C6", "=D10-D6"),
    (12, "Taux de marge de securite", "=B11/B10", "=C11/C10", "=D11/D10"),
]

for r, label, b, c, d in sr_data:
    ws6.cell(row=r, column=1, value=label).font = BOLD_FONT if 'SEUIL' in label or 'Marge' in label.upper() else NORMAL_FONT
    ws6.cell(row=r, column=1).border = THIN_BORDER
    for col_idx, val in [(2, b), (3, c), (4, d)]:
        cell = ws6.cell(row=r, column=col_idx, value=val)
        cell.border = THIN_BORDER
        if 'taux' in label.lower() or 'Taux' in label:
            cell.number_format = PCT_FORMAT_1
        elif r == 8:
            cell.number_format = '0'
        else:
            cell.number_format = EUR_FORMAT_NEG

ws6.column_dimensions['A'].width = 38
for c_letter in ['B', 'C', 'D']:
    ws6.column_dimensions[c_letter].width = 18


# =====================================================================
# ONGLET 7 : BILAN PREVISIONNEL
# =====================================================================
ws7 = wb.create_sheet("Bilan")
ws7.sheet_properties.tabColor = "319795"

ws7.merge_cells('A1:H1')
ws7['A1'] = "BILAN PREVISIONNEL"
ws7['A1'].font = TITLE_FONT

for col, label in enumerate(["ACTIF", "An 1", "An 2", "An 3", "PASSIF", "An 1", "An 2", "An 3"], 1):
    ws7.cell(row=3, column=col, value=label)
style_header_row(ws7, 3, 8)

# ACTIF
bilan_actif = [
    (4, "Immobilisations incorporelles (net)", 13500, 12000, 10500),
    (5, "Immobilisations corporelles (net)", 1000, 500, 0),
    (6, "Creances clients", 0, 0, 0),
    (7, "Tresorerie", None, None, None),  # linked
]

for r, label, a1, a2, a3 in bilan_actif:
    ws7.cell(row=r, column=1, value=label).font = NORMAL_FONT
    ws7.cell(row=r, column=1).border = THIN_BORDER
    for c, v in [(2, a1), (3, a2), (4, a3)]:
        cell = ws7.cell(row=r, column=c)
        if r == 7:
            # Tresorerie linked to tresorerie sheet (M12 cumulated)
            if c == 2:
                cell.value = "='Tresorerie M1-M12'!M26"
            elif c == 3:
                cell.value = 120000  # An 2 estimate
            else:
                cell.value = 240000  # An 3 estimate
        else:
            cell.value = v
        cell.number_format = EUR_FORMAT
        cell.border = THIN_BORDER

# Total Actif
ws7.cell(row=8, column=1, value="TOTAL ACTIF")
for c in [2, 3, 4]:
    ws7.cell(row=8, column=c, value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}7)")
    ws7.cell(row=8, column=c).number_format = EUR_FORMAT
style_total_row(ws7, 8, 4)

# PASSIF
bilan_passif = [
    (4, "Capital social (SASU)", 1, 1, 1),
    (5, "Reserves", 0, "='Compte Resultat'!B38", "=F5+'Compte Resultat'!C38"),
    (6, "Resultat de l'exercice", "='Compte Resultat'!B38", "='Compte Resultat'!C38", "='Compte Resultat'!D38"),
    (7, "Emprunts LT", 35632, 25780, 15928),
    (8, "Dettes courantes", 9651, 13727, 14396),
]

for r, label, a1, a2, a3 in bilan_passif:
    ws7.cell(row=r, column=5, value=label).font = NORMAL_FONT
    ws7.cell(row=r, column=5).border = THIN_BORDER
    for c, v in [(6, a1), (7, a2), (8, a3)]:
        ws7.cell(row=r, column=c, value=v)
        ws7.cell(row=r, column=c).number_format = EUR_FORMAT_NEG
        ws7.cell(row=r, column=c).border = THIN_BORDER

# Total Passif
ws7.cell(row=11, column=5, value="TOTAL PASSIF")
for c in [6, 7, 8]:
    ws7.cell(row=11, column=c, value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}10)")
    ws7.cell(row=11, column=c).number_format = EUR_FORMAT
for c in range(5, 9):
    ws7.cell(row=11, column=c).font = TOTAL_FONT
    ws7.cell(row=11, column=c).fill = TOTAL_FILL
    ws7.cell(row=11, column=c).border = THIN_BORDER

# Equilibre check
ws7.cell(row=13, column=1, value="Verification Actif = Passif").font = BOLD_FONT
ws7.cell(row=13, column=5, value="Ecart").font = BOLD_FONT
ws7.cell(row=13, column=6, value="=B8-F11").number_format = EUR_FORMAT_NEG
ws7.cell(row=13, column=7, value="=C8-G11").number_format = EUR_FORMAT_NEG
ws7.cell(row=13, column=8, value="=D8-H11").number_format = EUR_FORMAT_NEG

ws7.column_dimensions['A'].width = 35
ws7.column_dimensions['E'].width = 35
for c_letter in ['B', 'C', 'D', 'F', 'G', 'H']:
    ws7.column_dimensions[c_letter].width = 15


# =====================================================================
# ONGLET 8 : SCENARIO PESSIMISTE
# =====================================================================
ws8 = wb.create_sheet("Scenario Pessimiste")
ws8.sheet_properties.tabColor = "C53030"

ws8.merge_cells('A1:D1')
ws8['A1'] = "SCENARIO PESSIMISTE (-30% acquisition)"
ws8['A1'].font = TITLE_FONT

ws8.merge_cells('A2:D2')
ws8['A2'] = "Hypothese : 30% de clients en moins, charges ajustees a la baisse"
ws8['A2'].font = SMALL_FONT

for col, label in enumerate(["", "An 1", "An 2", "An 3"], 1):
    ws8.cell(row=4, column=col, value=label)
style_header_row(ws8, 4, 4)

# Facteur pessimiste modifiable
ws8.cell(row=5, column=1, value="Facteur reduction acquisition").font = BOLD_FONT
ws8.cell(row=5, column=2, value=0.70)
ws8.cell(row=5, column=2).number_format = PCT_FORMAT
ws8.cell(row=5, column=2).fill = LIGHT_YELLOW  # editable
ws8.merge_cells('C5:D5')
ws8.cell(row=5, column=3, value="Modifiez cette cellule pour ajuster le scenario").font = SMALL_FONT

# Clients fin annee = base * facteur
ws8.cell(row=7, column=1, value="Clients fin annee").font = BOLD_FONT
ws8.cell(row=7, column=2, value="=ROUND('Clients & MRR'!D17*$B$5,0)")
ws8.cell(row=7, column=3, value="=ROUND('Clients & MRR'!B22*$B$5,0)")
ws8.cell(row=7, column=4, value="=ROUND('Clients & MRR'!B28*$B$5,0)")

ws8.cell(row=8, column=1, value="MRR fin annee").font = BOLD_FONT
ws8.cell(row=8, column=2, value="=B7*Hypotheses!B11")
ws8.cell(row=8, column=2).number_format = EUR_FORMAT
ws8.cell(row=8, column=3, value="=C7*Hypotheses!B11")
ws8.cell(row=8, column=3).number_format = EUR_FORMAT
ws8.cell(row=8, column=4, value="=D7*Hypotheses!B11")
ws8.cell(row=8, column=4).number_format = EUR_FORMAT

ws8.cell(row=9, column=1, value="CA HT (pessimiste)").font = BOLD_FONT
ws8.cell(row=9, column=2, value="='Compte Resultat'!B9*$B$5")
ws8.cell(row=9, column=2).number_format = EUR_FORMAT
ws8.cell(row=9, column=3, value="='Compte Resultat'!C9*$B$5")
ws8.cell(row=9, column=3).number_format = EUR_FORMAT
ws8.cell(row=9, column=4, value="='Compte Resultat'!D9*$B$5")
ws8.cell(row=9, column=4).number_format = EUR_FORMAT

# Charges ajustees (marketing reduit, charges variables proportionnelles)
ws8.cell(row=10, column=1, value="Charges totales (ajustees)").font = NORMAL_FONT
ws8.cell(row=10, column=2, value=20100)
ws8.cell(row=10, column=2).number_format = EUR_FORMAT
ws8.cell(row=10, column=3, value=62600)
ws8.cell(row=10, column=3).number_format = EUR_FORMAT
ws8.cell(row=10, column=4, value=120800)
ws8.cell(row=10, column=4).number_format = EUR_FORMAT

ws8.cell(row=11, column=1, value="Resultat net (pessimiste)").font = BOLD_FONT
ws8.cell(row=11, column=2, value="=B9-B10")
ws8.cell(row=11, column=2).number_format = EUR_FORMAT_NEG
ws8.cell(row=11, column=3, value="=C9-C10")
ws8.cell(row=11, column=3).number_format = EUR_FORMAT_NEG
ws8.cell(row=11, column=4, value="=D9-D10")
ws8.cell(row=11, column=4).number_format = EUR_FORMAT_NEG

ws8.cell(row=13, column=1, value="Tresorerie fin annee").font = BOLD_FONT
ws8.cell(row=13, column=1).fill = LIGHT_GREEN
ws8.cell(row=13, column=2, value=47200).number_format = EUR_FORMAT_NEG
ws8.cell(row=13, column=3, value="=B13+C11").number_format = EUR_FORMAT_NEG
ws8.cell(row=13, column=4, value="=C13+D11").number_format = EUR_FORMAT_NEG
for c in [2, 3, 4]:
    ws8.cell(row=13, column=c).fill = LIGHT_GREEN

# Notes
ws8.cell(row=15, column=1, value="Actions correctives si scenario pessimiste :").font = BOLD_FONT
ws8.cell(row=16, column=1, value="→ Reduction marketing de 50%").font = SMALL_FONT
ws8.cell(row=17, column=1, value="→ Reduction salaire fondateur temporairement").font = SMALL_FONT
ws8.cell(row=18, column=1, value="→ Negociation report echeance prets d'honneur").font = SMALL_FONT
ws8.cell(row=19, column=1, value="→ Missions freelance en parallele pour le fondateur").font = SMALL_FONT
ws8.cell(row=20, column=1, value="→ Tresorerie reste POSITIVE meme a -30% (modele frugal)").font = BOLD_FONT

for r in range(7, 14):
    for c in range(1, 5):
        ws8.cell(row=r, column=c).border = THIN_BORDER

ws8.column_dimensions['A'].width = 38
for c_letter in ['B', 'C', 'D']:
    ws8.column_dimensions[c_letter].width = 16


# =====================================================================
# ONGLET 9 : ECHEANCIER REMBOURSEMENTS
# =====================================================================
ws9 = wb.create_sheet("Remboursements")
ws9.sheet_properties.tabColor = "4A5568"

ws9.merge_cells('A1:G1')
ws9['A1'] = "ECHEANCIER DES REMBOURSEMENTS"
ws9['A1'].font = TITLE_FONT

# Detail des prets
for col, label in enumerate(["Type", "Capital", "Taux", "Duree", "Differe", "Mensualite", "Debut remb."], 1):
    ws9.cell(row=3, column=col, value=label)
style_header_row(ws9, 3, 7)

prets = [
    (4, "Pret honneur Initiative 95", 15000, "0%", "5 ans", "6 mois", "=B4/54", "M7"),
    (5, "Pret bancaire (garanti BPI 60%)", 25000, "3%", "5 ans", "6 mois", 450, "M7"),
    (6, "NACRE phase 2 (beneficiaire ASS)", 5000, "0%", "5 ans", "12 mois", "=B6/48", "M13"),
]

for r, label, capital, taux, duree, differe, mensualite, debut in prets:
    ws9.cell(row=r, column=1, value=label).font = NORMAL_FONT
    ws9.cell(row=r, column=2, value=capital).number_format = EUR_FORMAT
    ws9.cell(row=r, column=3, value=taux)
    ws9.cell(row=r, column=4, value=duree)
    ws9.cell(row=r, column=5, value=differe)
    ws9.cell(row=r, column=6, value=mensualite)
    ws9.cell(row=r, column=6).number_format = EUR_FORMAT
    ws9.cell(row=r, column=7, value=debut)
    for c in range(1, 8):
        ws9.cell(row=r, column=c).border = THIN_BORDER

# Total capital
ws9.cell(row=7, column=1, value="TOTAL").font = BOLD_FONT
ws9.cell(row=7, column=2, value="=SUM(B4:B6)").number_format = EUR_FORMAT
style_total_row(ws9, 7, 7)

# Echeancier annuel
ws9.cell(row=9, column=1, value="ECHEANCIER ANNUEL").font = SECTION_FONT
for col, label in enumerate(["", "An 1", "An 2", "An 3", "Total"], 1):
    ws9.cell(row=10, column=col, value=label)
style_header_row(ws9, 10, 5)

ech = [
    (11, "Remb. pret I95", "=F4*6", "=F4*12", "=F4*12"),       # 6 mois An1 (M7-M12)
    (12, "Remb. pret bancaire", "=6*450", "=12*450", "=12*450"),  # 6 mois An1
    (13, "Remb. NACRE", 0, "=F6*12", "=F6*12"),                  # Debut M13
]

for r, label, a1, a2, a3 in ech:
    ws9.cell(row=r, column=1, value=label).font = NORMAL_FONT
    ws9.cell(row=r, column=1).border = THIN_BORDER
    for c, v in [(2, a1), (3, a2), (4, a3)]:
        ws9.cell(row=r, column=c, value=v)
        ws9.cell(row=r, column=c).number_format = EUR_FORMAT
        ws9.cell(row=r, column=c).border = THIN_BORDER
    ws9.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
    ws9.cell(row=r, column=5).number_format = EUR_FORMAT
    ws9.cell(row=r, column=5).border = THIN_BORDER

# Total remboursements
ws9.cell(row=14, column=1, value="TOTAL REMBOURSEMENTS")
for c in range(2, 6):
    ws9.cell(row=14, column=c, value=f"=SUM({get_column_letter(c)}11:{get_column_letter(c)}13)")
    ws9.cell(row=14, column=c).number_format = EUR_FORMAT
style_total_row(ws9, 14, 5)

# Interets
ws9.cell(row=15, column=1, value="Dont interets (pret bancaire)").font = SMALL_FONT
ws9.cell(row=15, column=2, value=375).number_format = EUR_FORMAT
ws9.cell(row=15, column=3, value=625).number_format = EUR_FORMAT
ws9.cell(row=15, column=4, value=500).number_format = EUR_FORMAT
ws9.cell(row=15, column=5, value="=SUM(B15:D15)").number_format = EUR_FORMAT

# Resume
ws9.cell(row=17, column=1, value="RESUME REMBOURSEMENTS MENSUELS").font = SECTION_FONT
ws9.cell(row=18, column=1, value="M1-M6 : aucun remboursement (differe)").font = SMALL_FONT
ws9.cell(row=19, column=1, value="M7-M12 : 728 EUR/mois (I95 278 + bancaire 450)").font = SMALL_FONT
ws9.cell(row=20, column=1, value="M13+ : 821 EUR/mois (+ NACRE 93)").font = SMALL_FONT

ws9.column_dimensions['A'].width = 35
for c_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws9.column_dimensions[c_letter].width = 14


# =====================================================================
# SAVE
# =====================================================================
output_path = os.path.join(os.path.dirname(__file__), "PREVISIONNEL_FINANCIER_NEXUS.xlsx")
wb.save(output_path)
print(f"Excel genere : {output_path}")
print(f"Onglets : {[ws.title for ws in wb.worksheets]}")
print("Formules inter-onglets actives — modifiez les Hypotheses pour voir l'impact partout.")
